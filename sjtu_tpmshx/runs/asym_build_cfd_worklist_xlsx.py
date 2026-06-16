"""asym_build_cfd_worklist_xlsx.py — Phase 1 asym-porosity CFD worklist (water-cfd-raw style).

Mirrors the prior water-side correlation-CFD design (TPMS水_关联式拟合CFD工况 →
water-cfd-raw.xlsx) so the asymmetric (offset-δ) dP calibration reuses the SAME
proven domain + post-processing:

  domain  = [15mm straight inlet] + [1×3 offset-TPMS core] + [15mm straight outlet]
            cross-section 1 cell 5×5mm, lateral x,y = PERIODIC; straight channels =
            the void face cross-section extruded (no contraction).
  BC      = mass-flow inlet ṁ = ρ·Um·(ε_side·L²); pressure outlet.
  extract = p0..p3 at the 3 core-cell boundaries; dp_core = p0−p3 (developed
            friction, entrance excluded) → Darcy_f_core → (K, c_F) fit.
  per-side= void_A (A-side runs) and void_B (B-side runs) are SEPARATE domains;
            same straight channels, only the core void differs.

κ relative-ratio: run the symmetric r=1 case in the SAME recipe → κ(r)=X(r)/X(1)
cancels every recipe artifact (entrance/exit/mesh/turbulence). One offset cell
gives both channels (kr_A>1 large, kr_B<1 small).

Output: runs/_out/asym_cfd/asym_cfd_worklist.xlsx
Usage:  python -u runs/asym_build_cfd_worklist_xlsx.py
"""
import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XLSide
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from solvers.tpms_geometry import _phi_grid, _C_from_tL
from solvers.asym_geometry import eps_sides, a0_sides_mc, dh_sides, percolates_z

# ── geometry design (locked) ─────────────────────────────────────
N = 128
L_mm = 5.0
t_mm = 0.4
TPMS = ["Diamond", "Gyroid"]
SPLITS = [1.0, 1.5, 2.0, 2.5, 3.0, 3.5]

# ── domain (mirrors water-cfd-raw) ───────────────────────────────
PERIOD_MM = 5.0        # one cell
N_CORE = 3             # 1×3 core
CORE_LEN_MM = 15.0     # = N_CORE × period
INLET_LEN_MM = 15.0    # straight, void-face extrude
OUTLET_LEN_MM = 15.0
LATERAL_BC = "periodic"

# ── per-side Re sweep (first 2 = low-Re Darcy-pin, rest = main) ───
RE_A = [100, 300, 1000, 3000, 6000, 12000]   # gas side; Re12000 @ smallest Dh → Um~98 m/s Mach<0.3
RE_B = [30, 80, 250, 700, 1800, 3000]        # liquid side
N_LOW = 2                                    # first N_LOW per side = Darcy-pin (low Re)

# ── fluid props (A=air ideal-gas @300K, B=water @325K reuse prior table) ──
# air @300K (denser than @350K → keeps Re12000 subsonic Mach<0.3); κ is geometry-
# only so the absolute Tref is immaterial to the calibration.
AIR = dict(name="air", Tref=300.0, P_MPa=0.101325, rho=1.1774, mu=1.846e-5,
           cp=1006.4, k=0.02624, Pr=0.708, Twall=400.0)
WATER = dict(name="water", Tref=325.0, P_MPa=0.101325, rho=987.11, mu=5.33e-4,
             cp=4180.9, k=0.643, Pr=3.4657, Twall=375.0)

OUT = Path(__file__).resolve().parents[1] / "runs" / "_out" / "asym_cfd"
XLSX = OUT / "asym_cfd_worklist.xlsx"

# ── styles ───────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ANCHOR_FILL = PatternFill("solid", fgColor="E2EFDA")   # symmetric r=1
FILL_FILL = PatternFill("solid", fgColor="FFF2CC")     # CFD-to-fill
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
THIN = XLSide(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")


def _delta_for_split(phi, C, target, n=6000):
    if target <= 1.0:
        return 0.0
    for d in np.linspace(0.0, float(np.abs(phi).max()), n):
        eA, eB, _ = eps_sides(phi, C, d)
        if eB <= 1e-9:
            break
        if eA / eB >= target:
            return float(d)
    return None


def _geom():
    """Per (tpms, split) geometry rows."""
    rows = []
    for tpms in TPMS:
        phi = _phi_grid(tpms, N)
        L_m = L_mm / 1000.0
        C = _C_from_tL(tpms, t_mm / L_mm)
        for r in SPLITS:
            d = _delta_for_split(phi, C, r)
            if d is None:
                continue
            eA, eB, eps = eps_sides(phi, C, d)
            e_sym = 0.5 * eps
            A0A, A0B = a0_sides_mc(phi, C, d, L_m, N)
            DhA, DhB = dh_sides(phi, C, d, L_m, N, mc=True)
            rows.append(dict(
                tpms=tpms, C=round(C, 4), delta=round(d, 4),
                phi_lo=round(d - C, 4), phi_hi=round(d + C, 4),
                split_r=r, eps_A=round(eA, 4), eps_B=round(eB, 4),
                eps_sym=round(e_sym, 4), solid=round(1 - eps, 4),
                kr_A=round(eA / e_sym, 4), kr_B=round(eB / e_sym, 4),
                A0_A=round(A0A, 1), A0_B=round(A0B, 1),
                Dh_A_mm=round(DhA * 1e3, 4), Dh_B_mm=round(DhB * 1e3, 4),
                pA=percolates_z(phi < (d - C)), pB=percolates_z(phi > (d + C)),
            ))
    return rows


def _hdr(ws, headers, row=1):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR; c.border = BORDER


def _widths(ws, ws_cols):
    for j, w in enumerate(ws_cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def build():
    OUT.mkdir(parents=True, exist_ok=True)
    geom = _geom()
    wb = Workbook()

    # ── README ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 118
    lines = [
        ("非对称孔隙率 Phase 1 — CFD 工况表（仿 water-cfd-raw 设计）", TITLE_FONT),
        ("", None),
        ("目标: per-side Darcy-Forchheimer (K, c_F) 的相对修正 κ_dP(r)。κ(r)=X(r)/X(r=1) 同 recipe 自比 → 抵消 entrance/exit/mesh/湍流。", None),
        ("", None),
        ("【流道几何（锁定）】", Font(bold=True, size=12)),
        (f"  域 = [{INLET_LEN_MM:g}mm 直通道进口] + [1×{N_CORE} offset-TPMS 核心 = {CORE_LEN_MM:g}mm] + [{OUTLET_LEN_MM:g}mm 直通道出口]，总 {INLET_LEN_MM+CORE_LEN_MM+OUTLET_LEN_MM:g}mm", None),
        (f"  横截面 = 1 胞 {L_mm:g}×{L_mm:g}mm；侧向 x,y = {LATERAL_BC}", None),
        ("  直通道 = 核心端面 void 截面直拉伸（无收缩，侧向也周期）", None),
        ("  进口 = 质量流量 ṁ = ρ·Um·(ε_side·L²)；出口 = 压力出口", None),
        (f"  压力面 p0..p3 在 {N_CORE} 核心胞边界（绝对 z = {INLET_LEN_MM:g},{INLET_LEN_MM+PERIOD_MM:g},{INLET_LEN_MM+2*PERIOD_MM:g},{INLET_LEN_MM+3*PERIOD_MM:g} mm）", None),
        ("  dp_core = p0 − p3（跨 3 核心胞，developed friction，入口已排）", None),
        ("", None),
        ("【offset-TPMS 核心方程（per-case 换）】 φ_族 见 lattice 列；固体壁 phi_lo≤φ≤phi_hi（worklist 列）；void_A={φ<phi_lo}(大/气) / void_B={φ>phi_hi}(小/液)。", None),
        ("  ⚠ nTop 建好先验体积分数 = eps_side（worklist 列）再 mesh。per-side: void_A 跑 A 行、void_B 跑 B 行，各一套（直通道相同，只核心换 void）。", None),
        ("", None),
        ("【物性】 A 侧 = air 可压 ideal-gas ρ(P,T) @Tref=350K；B 侧 = water @Tref=325K（复用水物性表）。κ 与流体无关，物性只定 Um/ṁ。", None),
        ("", None),
        ("【流程】 nTop 建 24 域(12 case×2 side) → Fluent 跑 worklist(Um/ṁ 已给) → 填黄列 p0..p3 → dp_core=p0−p3 → 每 (case,side) 4-6 Re 拟", None),
        ("  |ΔP|_core/L_core = (μ/K)·Um + c_F·ρ·Um²  → (K,c_F) → results sheet → python -m df_surrogate.ingest_cfd_kappa（先改成 κ(r)=X(r)/X(1)）。", None),
        ("", None),
        ("黄列 = 待 Fluent 填。绿行 = 对称锚 r=1。', '主拟合'=操作 Re 区(4 点); 低 Re 钉 Darcy K。", None),
    ]
    for i, (txt, font) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        if font:
            c.font = font

    # ── geom_cases (geometry reference) ──────────────────────────
    ws = wb.create_sheet("geom_cases")
    cols = ["lattice", "split_r", "C", "delta", "phi_lo", "phi_hi",
            "eps_A", "eps_B", "eps_sym", "solid", "kr_A", "kr_B",
            "Dh_A_mm", "Dh_B_mm", "A0_A_1/m", "A0_B_1/m", "conn", "ntop_label"]
    _hdr(ws, cols)
    for i, g in enumerate(geom, 2):
        conn = "OK" if (g["pA"] and g["pB"]) else ("cut-B" if not g["pB"] else "cut-A")
        vals = [g["tpms"], g["split_r"], g["C"], g["delta"], g["phi_lo"], g["phi_hi"],
                g["eps_A"], g["eps_B"], g["eps_sym"], g["solid"], g["kr_A"], g["kr_B"],
                g["Dh_A_mm"], g["Dh_B_mm"], g["A0_A"], g["A0_B"], conn,
                f"{g['tpms']}_L{int(L_mm)}_t{t_mm}_r{g['split_r']:g}"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v); c.border = BORDER; c.alignment = CTR
            if g["split_r"] == 1.0:
                c.fill = ANCHOR_FILL
    _widths(ws, [8, 7, 7, 7, 8, 8, 7, 7, 8, 7, 7, 7, 9, 9, 10, 10, 7, 24])
    ws.freeze_panes = "A2"

    # ── cfd_worklist (one row per case×side×Re; water-cfd-raw style) ──
    ws = wb.create_sheet("cfd_worklist")
    cols = ["case_id", "纳入主拟合", "lattice", "split_r", "side", "cell_mm",
            "delta", "phi_lo", "phi_hi", "eps_side", "eps_sym", "kr",
            "Dh_mm", "D/L", "Re", "fluid", "Tref_K", "P_MPa_abs",
            "rho", "mu", "cp", "k_cond", "Pr", "Pr^(1/3)",
            "Um_m_s", "mdot_kg_s", "Twall_K",
            "period_mm", "core_mm", "n_core", "inlet_mm", "outlet_mm", "lateral_BC",
            "p0_Pa(FILL)", "p1_Pa(FILL)", "p2_Pa(FILL)", "p3_Pa(FILL)",
            "dp_core_Pa(FILL)", "Darcy_f(FILL)", "note"]
    _hdr(ws, cols)
    L_m = L_mm / 1000.0
    rr = 2
    for g in geom:
        for side, eps_side, kr, Dh_mm, re_list, fl in (
            ("A", g["eps_A"], g["kr_A"], g["Dh_A_mm"], RE_A, AIR),
            ("B", g["eps_B"], g["kr_B"], g["Dh_B_mm"], RE_B, WATER),
        ):
            Dh_m = Dh_mm / 1e3
            for idx, re in enumerate(re_list):
                Um = re * fl["mu"] / (fl["rho"] * Dh_m) if Dh_m > 0 else 0.0
                mdot = fl["rho"] * Um * (eps_side * L_m ** 2)
                is_main = idx >= N_LOW
                anchor = (g["split_r"] == 1.0)
                cid = f"asym{g['tpms'][0]}_r{g['split_r']:g}_{side}_Re{re}"
                vals = [cid, is_main, g["tpms"], g["split_r"], side, L_mm,
                        g["delta"], g["phi_lo"], g["phi_hi"], eps_side, g["eps_sym"], kr,
                        Dh_mm, round(Dh_mm / L_mm, 4), re, fl["name"], fl["Tref"], fl["P_MPa"],
                        fl["rho"], fl["mu"], fl["cp"], fl["k"], fl["Pr"], round(fl["Pr"] ** (1 / 3), 4),
                        round(Um, 5), float("%.4g" % mdot), fl["Twall"],
                        PERIOD_MM, CORE_LEN_MM, N_CORE, INLET_LEN_MM, OUTLET_LEN_MM, LATERAL_BC,
                        None, None, None, None, None, None,
                        ("anchor r=1" if anchor else ("Darcy-pin" if not is_main else ""))]
                for j, v in enumerate(vals, 1):
                    c = ws.cell(row=rr, column=j, value=v); c.border = BORDER; c.alignment = CTR
                    if 34 <= j <= 39:                      # p0..p3, dp_core, Darcy_f
                        c.fill = FILL_FILL
                    elif anchor:
                        c.fill = ANCHOR_FILL
                rr += 1
    _widths(ws, [20, 9, 8, 7, 5, 7, 7, 8, 8, 8, 8, 7, 8, 7, 7, 7, 7, 9, 8, 9, 8, 8, 7, 8,
                 9, 11, 8, 8, 8, 7, 8, 8, 9, 12, 12, 12, 12, 14, 12, 11])
    ws.freeze_panes = "F2"
    n_runs = rr - 2

    # ── results_template (DF fit → ingest CSV) ───────────────────
    ws = wb.create_sheet("results_template")
    note = ws.cell(row=1, column=1,
                   value="每 (lattice,split,side) 对其 Re 的 (Um, dp_core/core_len) 拟 |ΔP|/L=(μ/K)Um+c_F·ρ·Um² → 填黄列 → 另存 CSV → ingest。")
    note.font = Font(italic=True, color="7F7F7F")
    cols = ["tpms", "L_mm", "t_mm", "eps_side", "eps_sym", "K_cfd(FILL)", "cF_cfd(FILL)"]
    _hdr(ws, cols, row=2)
    rr = 3
    for g in geom:
        for side, eps_side in (("A", g["eps_A"]), ("B", g["eps_B"])):
            vals = [g["tpms"], L_mm, t_mm, eps_side, g["eps_sym"], None, None]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=rr, column=j, value=v); c.border = BORDER; c.alignment = CTR
                if j in (6, 7):
                    c.fill = FILL_FILL
                elif g["split_r"] == 1.0:
                    c.fill = ANCHOR_FILL
            rr += 1
    _widths(ws, [9, 7, 6, 9, 9, 13, 13])
    ws.freeze_panes = "A3"

    wb.save(XLSX)
    print(f"[xlsx] {XLSX}")
    print(f"  geom_cases : {len(geom)} 几何")
    print(f"  cfd_worklist : {n_runs} 行 ({len(geom)} 几何 × ({len(RE_A)}A+{len(RE_B)}B Re))")
    print(f"  results : {len(geom)*2} per-side κ 点")


if __name__ == "__main__":
    build()
