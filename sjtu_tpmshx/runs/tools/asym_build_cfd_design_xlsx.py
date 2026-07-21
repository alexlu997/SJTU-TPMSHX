"""asym_build_cfd_design_xlsx.py — Phase 1 CFD design matrix → Excel.

Emits the per-side asymmetric-porosity (offset-isosurface δ) CFD calibration
matrix as a single styled .xlsx for ANSYS Fluent execution + κ ingest.

Design parameters (finalized 2026-06-15):
  topology : Diamond, Gyroid
  (L, t)   : 5 mm / 0.4 mm  → t/L = 0.08  (in water-cfd-raw grid; scale-invariant
             representative — C fixed by (topology, t/L), covers all same-ratio cells)
  split r  : ε_A/ε_B ∈ {1, 1.5, 2, 2.5, 3, 3.5}  (1 = symmetric κ=1 anchor)
             → δ back-solved per topology (fixed C, sweep δ)
  Re sweep : side A (large/gas)  {600, 2000, 6000, 12000, 20000}
             side B (small/liq)  {150, 400, 1000, 2000, 3000}
             (each side fit |ΔP|/L = (μ/K)u + c_F ρ u² → K, c_F)

κ method (relative-ratio): κ_X(r) = X_asym / X_sym, r = ε_side/ε_sym.
δ=0 → r=1 → κ=1 → bit-identical (zero regression). One offset cell gives BOTH
channels (A large r>1, B small r<1), so d↔1−d symmetry halves the geometry count.

Output: runs/_out/asym_cfd/asym_cfd_design_matrix.xlsx
Usage:  python -u runs/asym_build_cfd_design_xlsx.py
"""
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XLSide
from openpyxl.utils import get_column_letter

from sjtu_tpmshx.solvers.tpms_geometry import _phi_grid, _C_from_tL
from sjtu_tpmshx.solvers.asym_geometry import eps_sides, a0_sides_mc, dh_sides, percolates_z

# ── finalized design parameters ──────────────────────────────────
N = 128
L_mm = 5.0
t_mm = 0.4
TPMS = ["Diamond", "Gyroid"]
SPLITS = [1.0, 1.5, 2.0, 2.5, 3.0, 3.5]
RE_A = [600, 2000, 6000, 12000]   # side A: large channel (gas-like); cap Mach<0.3 (Um≲96 m/s)
RE_B = [150, 500, 1500, 3000]     # side B: small channel (liquid-like)
# nominal fluid props for inlet-velocity convenience (κ is geometry-only;
# props cancel in the ratio — Um here only seeds the Fluent BC).
AIR = dict(name="air",   rho=1.2,   mu=1.85e-5)
WATER = dict(name="water", rho=992.0, mu=6.5e-4)

OUT = Path(__file__).resolve().parents[2] / "runs" / "_out" / "asym_cfd"
XLSX = OUT / "asym_cfd_design_matrix.xlsx"

# ── styles ───────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
ANCHOR_FILL = PatternFill("solid", fgColor="E2EFDA")   # symmetric r=1 rows
FILL_FILL = PatternFill("solid", fgColor="FFF2CC")     # cells to fill from CFD
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
THIN = XLSide(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")


def _delta_for_split(phi, C, target, n=6000):
    """Smallest δ≥0 whose ε_A/ε_B ≥ target (fixed C). target≤1 → δ=0."""
    if target <= 1.0:
        return 0.0
    for d in np.linspace(0.0, float(np.abs(phi).max()), n):
        eA, eB, _ = eps_sides(phi, C, d)
        if eB <= 1e-9:
            break
        if eA / eB >= target:
            return float(d)
    return None


def _geom():
    """Compute per (tpms, split) geometry. Returns list of dict rows."""
    rows = []
    for tpms in TPMS:
        phi = _phi_grid(tpms, N)
        L_m = L_mm / 1000.0
        C = _C_from_tL(tpms, t_mm / L_mm)
        for r in SPLITS:
            d = _delta_for_split(phi, C, r)
            if d is None:
                print(f"[warn] {tpms} r={r} unreachable (pinch); skipped")
                continue
            eA, eB, eps = eps_sides(phi, C, d)
            e_sym = 0.5 * eps
            A0A, A0B = a0_sides_mc(phi, C, d, L_m, N)
            DhA, DhB = dh_sides(phi, C, d, L_m, N, mc=True)
            pA = percolates_z(phi < (d - C))
            pB = percolates_z(phi > (d + C))
            rows.append(dict(
                tpms=tpms, L_mm=L_mm, t_mm=t_mm, tL=round(t_mm / L_mm, 4),
                C=round(C, 4), delta=round(d, 4),
                phi_lo=round(d - C, 4), phi_hi=round(d + C, 4),
                solid=round(1.0 - eps, 4), split_r=r,
                eps_A=round(eA, 4), eps_B=round(eB, 4), eps_sym=round(e_sym, 4),
                kr_A=round(eA / e_sym, 4), kr_B=round(eB / e_sym, 4),
                A0_A=round(A0A, 1), A0_B=round(A0B, 1),
                Dh_A_mm=round(DhA * 1e3, 4), Dh_B_mm=round(DhB * 1e3, 4),
                conn=("OK" if (pA and pB) else ("cut-B" if not pB else "cut-A")),
                label=f"{tpms}_L{int(L_mm)}_t{t_mm}_r{r:g}",
            ))
    return rows


def _write_header(ws, headers, row=1):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR; c.border = BORDER


def _autosize(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def build():
    OUT.mkdir(parents=True, exist_ok=True)
    geom = _geom()
    wb = Workbook()

    # ── Sheet 1: README ──────────────────────────────────────────
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 110
    lines = [
        ("非对称孔隙率 Phase 1 — CFD 标定设计矩阵", TITLE_FONT),
        ("", None),
        ("目标: 标定 per-side Darcy-Forchheimer (K, c_F) 的相对修正 κ(r)，给 ε_A≠ε_B 偏移几何。", None),
        ("方法 = 相对比值: κ_X(r) = X_asym / X_sym,  r = ε_side/ε_sym。X_sym = 现有对称 baseline (predict_K_cF)。", None),
        ("比值抵消 sym/asym 共有 provenance (网格/湍流/粗糙) → 只留几何偏移效应。δ=0→r=1→κ=1→零回归。", None),
        ("", None),
        ("【设计参数】", Font(bold=True, size=12)),
        ("  topology : Diamond + Gyroid", None),
        (f"  (L, t)   : {L_mm} mm / {t_mm} mm  →  t/L = {t_mm/L_mm:g}  (在 water-cfd-raw 网格内)", None),
        ("            C 由 (topology, t/L) 定死 → 尺度不变, 此点代表所有同比例 (L,t)。", None),
        ("  split r  : ε_A/ε_B ∈ {1, 1.5, 2, 2.5, 3, 3.5}  (1 = 对称 κ=1 锚)", None),
        ("            固定 C, 反解 δ 命中 r (每族 δ 不同, 见 geom_cases)。", None),
        (f"  Re sweep : 侧A(大/气) {RE_A}", None),
        (f"             侧B(小/液) {RE_B}", None),
        ("            每侧拟 |ΔP|/L = (μ/K)·u + c_F·ρ·u² → 出该侧 (K, c_F)。", None),
        ("", None),
        ("【一个偏移胞元给两侧】 d↔1−d 对称: void_A(+δ)≅void_B(−δ)。一个 δ 的两通道 = κ-r>1 (大) + κ-r<1 (小) 两点。", None),
        ("", None),
        ("【几何 = nTop】 偏移-TPMS 胞元在 nTopology 建 (隐式建模, 原生 watertight + 周期)。", None),
        ("  nTop 输入/案: 等值面直接 δ−C≤φ≤δ+C 用 phi_lo/phi_hi 列; 或 thickness+offset 控, 用 eps_A/eps_B/solid 列做收敛-不变验证。", None),
        ("  ⚠ 先验 nTop 的 Gyroid/Diamond 振幅+iso 约定 == 本表 φ (range ±1.5): nTop 建出的 ε 须匹配 geom_cases ε 再 mesh。", None),
        ("  两股流体 = 同一道壁的互补 void; nTop 各输出 void_A/void_B 流体域 → Fluent。本仓 STL 导出 (asym_export_*) 已退役。", None),
        ("", None),
        ("【Fluent 设置】 压力基稳态 SIMPLE + ρ=ρ(P,T) ideal-gas + 流向平移周期 BC + coupled wall (仅要 dP 可单流体+壁热流)。", None),
        ("", None),
        ("【流程】", Font(bold=True, size=12)),
        ("  1. geom_cases → nTop 建偏移-TPMS 周期胞元 (验 ε 匹配) → 导出流体域给 Fluent。", None),
        ("  2. cfd_runs: 每 (case, side, Re) 一行, Um 已给 → Fluent 跑 → 填黄色 dP_Pa 列。", None),
        ("  3. 每 (tpms, split, side) 对其 Re 的 (u, dP) 拟 DF → 得 (K_cfd, c_F_cfd)。", None),
        ("  4. 填 results_template 黄色 K_cfd/cF_cfd 列 → 另存 CSV → python -m df_surrogate.ingest_cfd_kappa <csv>。", None),
        ("  5. $env:TPMSHX_ASYM_KAPPA=1 → 跑 Shanghai 3D (δ=0→κ=1) 必复现 9.82/3.20 (零回归闸)。", None),
        ("", None),
        (f"  Um 名义流体 (仅给入口速度, κ 与流体无关): air ρ={AIR['rho']}/μ={AIR['mu']}, water ρ={WATER['rho']}/μ={WATER['mu']}。", None),
        ("  Um = Re·μ/(ρ·D_h)。黄色单元格 = 待 CFD 填。绿色行 = 对称锚 (r=1)。", None),
    ]
    for i, (txt, font) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        if font:
            c.font = font
        c.alignment = Alignment(vertical="center", wrap_text=False)

    # ── Sheet 2: geom_cases ──────────────────────────────────────
    ws = wb.create_sheet("geom_cases")
    cols = ["tpms", "L_mm", "t_mm", "t/L", "C", "delta", "phi_lo", "phi_hi",
            "split_r", "eps_A", "eps_B", "eps_sym", "solid", "kr_A", "kr_B",
            "A0_A_1/m", "A0_B_1/m", "Dh_A_mm", "Dh_B_mm", "conn", "ntop_label"]
    _write_header(ws, cols)
    for i, g in enumerate(geom, 2):
        vals = [g["tpms"], g["L_mm"], g["t_mm"], g["tL"], g["C"], g["delta"],
                g["phi_lo"], g["phi_hi"], g["split_r"], g["eps_A"], g["eps_B"],
                g["eps_sym"], g["solid"], g["kr_A"], g["kr_B"], g["A0_A"],
                g["A0_B"], g["Dh_A_mm"], g["Dh_B_mm"], g["conn"], g["label"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = BORDER; c.alignment = CTR
            if g["split_r"] == 1.0:
                c.fill = ANCHOR_FILL
    _autosize(ws, [9, 7, 6, 6, 7, 7, 8, 8, 8, 7, 7, 8, 7, 7, 7, 10, 10, 9, 9, 7, 26])
    ws.freeze_panes = "A2"

    # ── Sheet 3: cfd_runs (dP to fill) ───────────────────────────
    ws = wb.create_sheet("cfd_runs")
    cols = ["case_id", "tpms", "split_r", "side", "eps_side", "kr",
            "Dh_mm", "fluid", "Re", "Um_m_s", "dP_Pa (FILL)", "L_period_mm", "note"]
    _write_header(ws, cols)
    rr = 2
    for g in geom:
        for side, eps_side, kr, Dh_mm, re_list, fluid in (
            ("A", g["eps_A"], g["kr_A"], g["Dh_A_mm"], RE_A, AIR),
            ("B", g["eps_B"], g["kr_B"], g["Dh_B_mm"], RE_B, WATER),
        ):
            Dh_m = Dh_mm / 1e3
            for re in re_list:
                Um = re * fluid["mu"] / (fluid["rho"] * Dh_m) if Dh_m > 0 else 0.0
                cid = f"{g['tpms']}_r{g['split_r']:g}_{side}_Re{re}"
                vals = [cid, g["tpms"], g["split_r"], side, eps_side, kr,
                        Dh_mm, fluid["name"], re, round(Um, 4), None,
                        g["L_mm"], ("anchor r=1" if g["split_r"] == 1.0 else "")]
                for j, v in enumerate(vals, 1):
                    c = ws.cell(row=rr, column=j, value=v)
                    c.border = BORDER; c.alignment = CTR
                    if j == 11:                      # dP_Pa fill column
                        c.fill = FILL_FILL
                    elif g["split_r"] == 1.0:
                        c.fill = ANCHOR_FILL
                rr += 1
    _autosize(ws, [26, 9, 8, 5, 9, 6, 8, 7, 8, 9, 13, 12, 12])
    ws.freeze_panes = "A2"

    # ── Sheet 4: results_template (ingest CSV format) ────────────
    ws = wb.create_sheet("results_template")
    note = ws.cell(row=1, column=1,
                   value="拟合 DF 后填黄色 K_cfd / cF_cfd → 另存为 CSV (仅这 7 列) → ingest_cfd_kappa。对称 r=1 行: K_cfd/cF_cfd 应 ≈ baseline → κ≈1 (验证 CFD setup)。")
    note.font = Font(italic=True, color="7F7F7F")
    cols = ["tpms", "L_mm", "t_mm", "eps_side", "eps_sym", "K_cfd (FILL)", "cF_cfd (FILL)"]
    _write_header(ws, cols, row=2)
    rr = 3
    for g in geom:
        for side, eps_side in (("A", g["eps_A"]), ("B", g["eps_B"])):
            vals = [g["tpms"], g["L_mm"], g["t_mm"], eps_side, g["eps_sym"], None, None]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=rr, column=j, value=v)
                c.border = BORDER; c.alignment = CTR
                if j in (6, 7):
                    c.fill = FILL_FILL
                elif g["split_r"] == 1.0:
                    c.fill = ANCHOR_FILL
            rr += 1
    _autosize(ws, [9, 7, 6, 9, 9, 14, 14])
    ws.freeze_panes = "A3"

    wb.save(XLSX)
    n_geom = len(geom)
    n_runs = sum(len(RE_A) + len(RE_B) for _ in geom)
    print(f"[xlsx] {XLSX}")
    print(f"  geom_cases : {n_geom} 几何 (2 topo × {len(SPLITS)} split)")
    print(f"  cfd_runs   : {n_runs} 行 (每几何 {len(RE_A)}A+{len(RE_B)}B Re)")
    print(f"  results    : {n_geom*2} per-side κ 点")
    print("  → 填 dP / (K,c_F) 后 ingest → 过 Shanghai 闸")


if __name__ == "__main__":
    build()
