"""merge_cfd3_to_legacy.py — merge per-geometry CFD3 sheets into legacy
"试验记录表_整理版" structure (Diamond_汇总 + Gyroid_汇总 single-sheet
combined layout) with the updated Nu values.

Layout reproduced:
  row 0  — column headers (49 cols, labels from legacy file)
  per geometry block:
    banner row '==== {LABEL} (L=Lmm, t=tmm, ε=...) ===='
    data rows: [LABEL, L, t, Re, D, Re_corr, A, T_inlet, P_inlet, mu, Cp,
                 m, rho, u, ..., Nu, ..., 摩擦压损]
    optional trailing '总密度' row (skipped — not in CFD3 source)

Output: data/raw_data/试验记录表_整理版_v2.xlsx (does NOT overwrite legacy
to allow diff). Sheets Diamond_汇总, Gyroid_汇总, 边界效应系数 (copied
verbatim from legacy since CFD3 lacks it under the same name).
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
import numpy as np

_PROJECT = Path(__file__).resolve().parent.parent.parent
SRC_LEGACY = _PROJECT / 'data' / 'raw_data' / '试验记录表_整理版.xlsx'
SRC_CFD3   = _PROJECT / 'data' / 'raw_data' / '试验记录表_CFD3.xlsx'
DST        = _PROJECT / 'data' / 'raw_data' / '试验记录表_整理版_v2.xlsx'

# Legacy 49-col layout (matches Diamond_汇总 row 0):
LEGACY_HEADERS = [
    '编号', 'L/mm', 't/mm', 'Re', 'D/mm', '尺寸距形修正Re', 'A/mm2',
    'T_inlet/℃', 'P_inlet/Pa', '动力粘度Pas', 'Cp/J/kgK', 'm/kg/s',
    '密度/kg/m3', '速度/m/s', '标准流量SLM', '实际流量SLM',
    '电流1/A', '电压1/V', '电流2/A', '电压2/V',
    '功率1/W', '功率2/W', '总加热功率/W', '出口温度/℃',
    '加热面热密度/W/m2', '散热/W/m2', '实际加热功率/W/m2',
    'CFD质量流量/kg/s', '实际出口温度/℃', 'CFD出口温度/K',
    'CFD_inletT', 'CFD_outletT', 'CFD_TPMST',
    'CFD_solidwall3T', 'CFD_solidwall4T', 'CFD_inletP',
    'Area_TPMS', 'Area_Solidwall3', 'Area_Solidwall4',
    'h/W/m2K', 'Nu', 'Pressureloss', 'Inlet_Pressureloss',
    'Pressureloss_TPMS', 'P_Exp/P_CFD', 'f', '转折f', '摩擦压损',
    'G (千克每平方米每秒)',
]
N_LEGACY_COLS = 49

# CFD3 sheet has 45 cols matching legacy cols 3..47 (Re ... 摩擦压损).
# Legacy col 0 = label, 1 = L/mm, 2 = t/mm, 48 = G (千克每平方米每秒) —
# G computed = m / A (kg/s / m² = kg/(m²·s)).


def _label(tpms: str, L_mm: float, t_mm: float) -> str:
    """Build legacy label e.g. 'D_8_03', 'G_6_05'."""
    prefix = 'D' if tpms == 'Diamond' else 'G'
    L_int = int(round(L_mm))
    t_int = int(round(t_mm * 10))   # 0.3 → 03, 0.4 → 04
    return f"{prefix}_{L_int}_{t_int:02d}"


def _epsilon_for(tpms: str, L_mm: float, t_mm: float) -> float:
    """Look up ε for banner string (matches legacy banner)."""
    sys.path.insert(0, str(_PROJECT / 'sjtu_tpmshx'))
    from solvers.tpms_calc import geometry as tpms_geometry
    g = tpms_geometry(tpms, float(L_mm), float(t_mm), 16.0)
    return float(g['epsilon'])


def _build_combined(tpms: str) -> list[list]:
    """Read all CFD3 per-geometry sheets for a TPMS type, return rows
    (list of lists) shaped (N, 49) including header + banners."""
    xls = pd.ExcelFile(SRC_CFD3, engine='openpyxl')
    out: list[list] = []
    out.append(LEGACY_HEADERS)
    # Iterate geometries in legacy order: L=8, 6, 5, 4 × t=0.3, 0.4, 0.5
    for L in (8, 6, 5, 4):
        for t10 in (3, 4, 5):
            sheet = f'{tpms}{L}_{t10}'
            if sheet not in xls.sheet_names:
                continue
            df = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl')
            if df.shape[1] < 45:
                print(f"  skip {sheet}: only {df.shape[1]} cols")
                continue
            t_mm = t10 / 10.0
            label = _label(tpms, L, t_mm)
            eps_full = _epsilon_for(tpms, L, t_mm)
            # Banner row
            banner = [None] * N_LEGACY_COLS
            banner[0] = (
                f"==== {label}  (L={L}mm, t={t_mm}mm, ε={eps_full:.3f}) ===="
            )
            out.append(banner)
            # Data rows
            for ridx in range(len(df)):
                try:
                    Re = df.iloc[ridx, 0]
                    if pd.isna(Re):
                        continue
                except Exception:
                    continue
                row = [None] * N_LEGACY_COLS
                row[0] = label
                row[1] = L
                row[2] = t_mm
                # CFD3 cols 0..44 → legacy cols 3..47
                for j in range(45):
                    val = df.iloc[ridx, j]
                    if pd.isna(val):
                        row[3 + j] = None
                    else:
                        row[3 + j] = val
                # G (col 48) = m_kg/s / A_mm2·1e-6 if both present
                m = row[11]; A_mm2 = row[6]
                if m is not None and A_mm2 is not None and A_mm2 != 0:
                    try:
                        row[48] = float(m) / (float(A_mm2) * 1e-6)
                    except Exception:
                        row[48] = None
                out.append(row)
    return out


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass
    print(f"Building merged Excel from {SRC_CFD3.name} → {DST.name}")
    diamond_rows = _build_combined('Diamond')
    gyroid_rows  = _build_combined('Gyroid')
    print(f"  Diamond_汇总: {len(diamond_rows)} rows (incl banners + header)")
    print(f"  Gyroid_汇总:  {len(gyroid_rows)} rows")

    # Carry forward 边界效应系数 verbatim from legacy
    legacy = pd.ExcelFile(SRC_LEGACY, engine='openpyxl')
    aux_sheet = '边界效应系数'
    aux_df = None
    for s in legacy.sheet_names:
        if s.endswith('边界效应系数') or '边界' in s:
            aux_df = pd.read_excel(legacy, sheet_name=s, engine='openpyxl',
                                    header=None)
            break

    # Write via openpyxl directly so banner rows (with None pad cells)
    # are not coerced into all-NaN rows by pandas.to_excel.
    # Banner cells starting with `=` would be parsed as formulas — force
    # text type to prevent Excel "removed records: 公式" recovery on open.
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    wb = Workbook()
    wb.remove(wb.active)

    def _write_sheet(name: str, rows: list[list]) -> None:
        ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                if val is None:
                    continue
                if isinstance(val, str) and val.startswith('='):
                    # Force text — prefix apostrophe Excel-style + explicit
                    # data_type so openpyxl writes <c t="str"> not <c t="e">.
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.data_type = 's'
                else:
                    ws.cell(row=r_idx, column=c_idx, value=val)

    _write_sheet('Diamond_汇总', diamond_rows)
    _write_sheet('Gyroid_汇总', gyroid_rows)
    if aux_df is not None:
        ws = wb.create_sheet(aux_sheet)
        for _, row in aux_df.iterrows():
            ws.append([None if pd.isna(v) else v for v in row.tolist()])
        print(f"  copied {aux_sheet} from legacy ({aux_df.shape})")
    wb.save(DST)

    print(f"\nWrote {DST}")
    print(f"Size: {DST.stat().st_size:,} bytes")


if __name__ == '__main__':
    main()
