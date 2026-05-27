"""设计工况 schema 与多行多列 loader (xlsx / csv)。"""
from __future__ import annotations
from dataclasses import dataclass
import os
import csv as _csv
import openpyxl

@dataclass
class DesignCase:
    case: int
    hot_fluid: str; T_in_h: float; P_in_h: float; mdot_h: float
    cold_fluid: str; T_in_c: float; P_in_c: float; mdot_c: float
    Q: float | None            # 换热量 [W] (Q 与 dT 至少一)
    dPlim_h: float; dPlim_c: float   # 分数 (ΔP/P_in)
    dT: float | None = None    # 热侧温降 [K] (与 Q 二选一, 优先)

# 必备基础列; duty 列 (Q_kW / dT_h_K) 至少出现一个
_BASE = ["case","hot_fluid","T_in_h_K","P_in_h_kPa","mdot_h",
         "cold_fluid","T_in_c_K","P_in_c_kPa","mdot_c","dPlim_h","dPlim_c"]

def _blank(v):
    """Return None if v is None or an empty/whitespace string, else return v."""
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v

def _row_to_case(get) -> "DesignCase | None":
    """get(name) -> raw value (None means absent/empty). case 空 → 返回 None(跳过)。"""
    if _blank(get("case")) is None:
        return None
    Qv  = _blank(get("Q_kW"))
    dTv = _blank(get("dT_h_K"))
    if Qv is None and dTv is None:
        raise ValueError(f"工况 {get('case')}: Q 与 dT 均空")
    return DesignCase(
        case=int(float(get("case"))),
        hot_fluid=str(get("hot_fluid")).strip().lower(),
        T_in_h=float(get("T_in_h_K")), P_in_h=float(get("P_in_h_kPa"))*1e3,
        mdot_h=float(get("mdot_h")),
        cold_fluid=str(get("cold_fluid")).strip().lower(),
        T_in_c=float(get("T_in_c_K")), P_in_c=float(get("P_in_c_kPa"))*1e3,
        mdot_c=float(get("mdot_c")),
        Q=(float(Qv)*1e3 if Qv is not None else None),
        dPlim_h=float(get("dPlim_h")), dPlim_c=float(get("dPlim_c")),
        dT=(float(dTv) if dTv is not None else None))

def _check_duty_cols(header):
    if "Q_kW" not in header and "dT_h_K" not in header:
        raise ValueError("缺 duty 列: 需 Q_kW 或 dT_h_K 至少一个")

def _load_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    _check_duty_cols(header)
    idx = {name: header.index(name) for name in _BASE}
    pos = {**idx,
           "Q_kW":   header.index("Q_kW")   if "Q_kW"   in header else None,
           "dT_h_K": header.index("dT_h_K") if "dT_h_K" in header else None}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def get(n, _row=row, _pos=pos):
            j = _pos[n]; return _row[j] if j is not None else None
        c = _row_to_case(get)
        if c is not None:
            out.append(c)
    return out

def _load_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(_csv.DictReader(f))
    if not rows:
        return []
    header = list(rows[0].keys())
    for col in _BASE:
        if col not in header:
            raise ValueError(f"缺列: {col}")
    _check_duty_cols(header)
    out = []
    for r in rows:
        def get(n, _r=r): return _r.get(n)
        c = _row_to_case(get)
        if c is not None:
            out.append(c)
    return out

def load_cases(path: str) -> list[DesignCase]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _load_csv(path)
    return _load_xlsx(path)
