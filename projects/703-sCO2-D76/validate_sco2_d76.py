"""Gate A — sCO2 Nu closure vs D-7-6 experiment (lumped dual-Nu ε-NTU).

Phase A validation (2026-06-26). For each D-7-6 GOLD case (clean, large
ΔT_streams, |energy-balance|<5%), drive the solver's sCO2 heat-transfer
closure (compute(fluid_type='sco2') -> H_sf both sides), assemble UA, run a
counterflow ε-NTU, and compare the predicted duty Q to the measured
enthalpy-based Q (= ṁ·Δh). Both streams are sCO2 (symmetric dual-Nu).

This is the lumped-closure gate (mirrors validate_shanghai_lumped_dual_nu.py
for air). It tests the Nu CORRELATION, not the field SIMPLE/LTNE solver
(Phase A++); the ε-NTU duty is a near-independent check of the construced-wall
Nu reduction the experiment used.

Geometry/conditions read from
  data/raw_data/D-7-6-sCO2/D-7-6实验数据-V1.xlsx
Run:  python projects/703-sCO2-D76/validate_sco2_d76.py
Gate: max per-case |Q error| < 15 %.
"""

# ⚠ HISTORY — 2026-07-15: solver sCO2 closures switched to SMOOTH-WALL
# unit-cell CFD fits, this gate suspended (errors expected ~1.7× on Nu until
# a roughness anchor lands). 2026-07-22 (candidate D · D-2sc-3/4): the
# experimental anchors LANDED — gamma_nu_sco2 (Nu, in compute()'s chain) and
# gamma_f_sco2 (cF) — so this gate is re-armed: it now validates the
# CORRECTED closure end-to-end (GOLD-case hot/cold Re ≈ 9–14 k, inside the
# gamma_Nu window). Data path updated: the old D-7-6-sCO2/…V1.xlsx moved to
# the flat data/raw_data/D-7-6实验数据-sCO2.xlsx re-export — SAME dataset
# (cases 15/20/38 cross-checked to 6 decimals against sCO2-Experient.xlsx)
# with every column shifted one LEFT vs V1; the map below is header-guarded
# so a future re-export cannot silently shift again (the sCO2-Experient
# Gyroid-sheet "+2 offset" lesson).
import math
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.utils import column_index_from_string as ci

_HERE = Path(__file__).resolve()

from sjtu_tpmshx.solvers import fluid_props, tpms_calc      # noqa: E402

XLSX = (_HERE.parent.parent.parent / "data" / "raw_data"
        / "D-7-6实验数据-sCO2.xlsx")

GOLD_CASES = [15, 20, 21, 32, 37, 38]   # ΔT_streams>10 °C & |bal|<5 %
K_WALL = 16.0                            # solid conductivity [W/m·K] (steel spec)
GATE_PCT = 15.0

# Flat re-export column map (sheet 整理版, rows 3–53 = 51 cases), with the
# header substring each column MUST carry (guarded in _col_checked).
_COLMAP = {
    "case":   ("A", "序号"),
    "L_m":    ("B", "流道长度"),
    "A_flow": ("D", "流通截面积"),
    "A_ht":   ("E", "换热面积"),
    "mh":     ("F", "质量流量"),
    "Th":     ("G", "入口温度"),
    "Ph":     ("H", "入口压力"),
    "mc":     ("K", "质量流量"),
    "Tc":     ("L", "入口温度"),
    "Pc":     ("M", "入口压力"),
    "Qexp":   ("R", "换热量"),
}


def _col_checked(ws, key):
    L, must = _COLMAP[key]
    hdr = "".join(str(ws.cell(r, ci(L)).value or "") for r in (1, 2))
    if must not in hdr:
        raise RuntimeError(
            f"D-7-6 xlsx column {L} header {hdr!r} lacks {must!r} — the "
            f"re-export layout shifted again; re-derive _COLMAP.")
    return np.array([ws.cell(r, ci(L)).value for r in range(3, 54)], float)


def _eps_counterflow(NTU, Cr):
    if abs(1.0 - Cr) < 1e-9:
        return NTU / (1.0 + NTU)
    e = math.exp(-NTU * (1.0 - Cr))
    return (1.0 - e) / (1.0 - Cr * e)


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    L_m = _col_checked(ws, 'L_m')[0]        # 流道长度 0.182 m
    A_flow = _col_checked(ws, 'A_flow')[0]  # 截面流通面积 [m^2]
    A_ht = _col_checked(ws, 'A_ht')[0]      # 总换热面积 [m^2]
    L_cell_mm, wall_mm = 7.0, 0.6

    case = _col_checked(ws, 'case')
    mh, ThI, PhI = (_col_checked(ws, 'mh'), _col_checked(ws, 'Th'),
                    _col_checked(ws, 'Ph'))
    mc, TcI, PcI = (_col_checked(ws, 'mc'), _col_checked(ws, 'Tc'),
                    _col_checked(ws, 'Pc'))
    Qexp = _col_checked(ws, 'Qexp')         # 换热量 kW (ṁ·Δh, hot side)

    m = fluid_props.get('sco2')
    print(f"sCO2 D-7-6 Gate A — geometry L={L_m} m, A_ht={A_ht:.4f} m^2, "
          f"A_flow={A_flow:.2e} m^2, Diamond {L_cell_mm}/{wall_mm}")
    print(f"{'case':>4} {'Re_h':>7} {'Re_c':>7} {'Q_exp':>7} {'Q_pred':>7} "
          f"{'err%':>7}")

    errs = []
    for cidx in GOLD_CASES:
        i = int(np.where(case == cidx)[0][0])
        Th, Ph = ThI[i] + 273.15, PhI[i] * 1e6
        Tc, Pc = TcI[i] + 273.15, PcI[i] * 1e6
        # interstitial velocity each side
        uh = mh[i] / (m.rho(Th, Ph) * A_flow)
        uc = mc[i] / (m.rho(Tc, Pc) * A_flow)
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            rh = tpms_calc.compute('Diamond', L_cell_mm, wall_mm, uh, Th, Ph,
                                   K_WALL, fluid_type='sco2')
            rc = tpms_calc.compute('Diamond', L_cell_mm, wall_mm, uc, Tc, Pc,
                                   K_WALL, fluid_type='sco2')
        h_h, h_c = rh['H_sf'], rc['H_sf']
        # series resistances: conv_hot + wall + conv_cold
        R = 1.0 / (h_h * A_ht) + (wall_mm / 1000.0) / (K_WALL * A_ht) \
            + 1.0 / (h_c * A_ht)
        UA = 1.0 / R
        Ch = mh[i] * m.cp(Th, Ph)
        Cc = mc[i] * m.cp(Tc, Pc)
        Cmin, Cmax = min(Ch, Cc), max(Ch, Cc)
        eps = _eps_counterflow(UA / Cmin, Cmin / Cmax)
        Qpred = eps * Cmin * (Th - Tc) / 1000.0     # kW
        err = (Qpred - Qexp[i]) / Qexp[i] * 100.0
        errs.append(err)
        print(f"{cidx:>4} {rh['Re']:>7.0f} {rc['Re']:>7.0f} {Qexp[i]:>7.2f} "
              f"{Qpred:>7.2f} {err:>+7.1f}")

    errs = np.array(errs)
    rmsre = math.sqrt(np.mean(errs ** 2))
    mx = np.max(np.abs(errs))
    print(f"\nRMSRE = {rmsre:.1f}%  max|err| = {mx:.1f}%  "
          f"(bias {np.mean(errs):+.1f}%)")
    ok = mx < GATE_PCT
    print(f"GATE A ({GATE_PCT:.0f}% max): {'PASS' if ok else 'FAIL'}")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
